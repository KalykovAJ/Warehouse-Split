import os
import glob
from tkinter import Tk, filedialog
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
from collections import defaultdict


# =====================================================
# НАСТРОЙКА: путь к папке со справочниками
# =====================================================

CATALOGS_DIR = r"C:\Users\Пользователь\Desktop\Справочники"


# =====================================================
# ОЧИСТКА ТЕКСТА
# =====================================================

def clean_text(text):
    if pd.isna(text):
        return ""
    text = str(text)
    text = text.replace("\n", " ").replace("\r", " ").replace(chr(160), " ")
    return " ".join(text.lower().split())


# =====================================================
# ЗАГРУЗКА ВСЕХ СПРАВОЧНИКОВ
# Файлы: Справочник_*.xlsm / Справочник_*.xlsx
# Имя сети = название ПЕРВОГО листа файла
# Колонки определяются по заголовку (Наименование, Склад)
# =====================================================

def load_all_catalogs():
    if not os.path.isdir(CATALOGS_DIR):
        print(f"Ошибка: Папка справочников не найдена:\n{CATALOGS_DIR}")
        exit()

    found_files = (
        glob.glob(os.path.join(CATALOGS_DIR, "Справочник *.xlsm")) +
        glob.glob(os.path.join(CATALOGS_DIR, "Справочник *.xlsx"))
    )

    if not found_files:
        print(f"Ошибка: В папке не найдено ни одного файла Справочник_*.xlsm / .xlsx\n{CATALOGS_DIR}")
        exit()

    catalogs = {}       # { network_lower: { product_clean: warehouse } }
    network_names = {}  # { network_lower: original_name }

    for filepath in found_files:
        filename = os.path.basename(filepath)
        try:
            wb = load_workbook(filepath, data_only=True)
        except Exception as e:
            print(f"Ошибка при открытии '{filename}': {e}")
            continue

        # Берём только ПЕРВЫЙ лист
        ws = wb.worksheets[0]
        network_original = ws.title.strip()
        network_lower    = network_original.lower()

        if network_lower in catalogs:
            print(f"Предупреждение: Сеть '{network_original}' уже загружена, пропускаю дубль в '{filename}'.")
            wb.close()
            continue

        # Определяем колонки по заголовку первой строки
        headers = {}
        for cell in ws[1]:
            if cell.value:
                headers[clean_text(str(cell.value))] = cell.column - 1  # 0-based

        name_col      = next((i for h, i in headers.items() if "наименование" in h), 1)
        warehouse_col = next((i for h, i in headers.items() if "склад" in h), 8)

        product_dict = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= max(name_col, warehouse_col):
                continue
            product   = clean_text(row[name_col])
            warehouse = str(row[warehouse_col]).strip() if row[warehouse_col] else ""
            if product and warehouse:
                product_dict[product] = warehouse

        wb.close()

        if product_dict:
            catalogs[network_lower]      = product_dict
            network_names[network_lower] = network_original
        else:
            print(f"Предупреждение: Справочник '{filename}' (сеть '{network_original}') пустой, пропускаю.")

    print(f"[Справочники] Загружено сетей: {len(catalogs)} из {len(found_files)} файлов")
    for nl, nn in network_names.items():
        print(f"   ▪ {nn} ({len(catalogs[nl])} позиций)")

    return catalogs, network_names


# =====================================================
# ЗАГРУЗКА СПРАВОЧНИКОВ
# =====================================================

all_catalogs, all_network_names = load_all_catalogs()


# =====================================================
# ВЫБОР ОСНОВНОГО ФАЙЛА ОПЕРАТОРОМ
# =====================================================

root = Tk()
root.withdraw()

main_file = filedialog.askopenfilename(
    title="Выберите основной файл для обработки",
    filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")]
)

if not main_file:
    print("Действие отменено пользователем.")
    exit()


# =====================================================
# ЗАГРУЗКА ОСНОВНОГО ФАЙЛА
# =====================================================

wb = load_workbook(main_file, keep_vba=True)
not_found = []


# =====================================================
# СТИЛИ ОФОРМЛЕНИЯ
# =====================================================

title_font     = Font(name="Calibri", size=16, bold=True)
warehouse_font = Font(name="Calibri", size=15, bold=True)
header_font    = Font(name="Calibri", size=14, bold=True)
text_font      = Font(name="Calibri", size=13)

center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
left      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
thin_side = Side(style="thin", color="000000")
border    = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)


# =====================================================
# ОСНОВНОЙ ЦИКЛ ОБРАБОТКИ ЛИСТОВ
# =====================================================

processed_sheets = 0

for ws in wb.worksheets:
    sheet_title_lower = ws.title.strip().lower()

    matched_network = next(
        (net for net in all_catalogs if net in sheet_title_lower),
        None
    )

    if matched_network is None:
        continue

    processed_sheets += 1
    azs_name     = ws.title.upper()
    product_dict = all_catalogs[matched_network]

    data = []
    for row in range(2, ws.max_row + 1):
        num          = ws.cell(row=row, column=1).value
        product_name = ws.cell(row=row, column=2).value
        qty          = ws.cell(row=row, column=3).value

        clean_product = clean_text(product_name)
        if clean_product in ("", "товар", "наименование", "продукт"):
            continue

        warehouse = product_dict.get(clean_product, "НЕ НАЙДЕНО")
        if warehouse == "НЕ НАЙДЕНО":
            not_found.append((ws.title, product_name))

        data.append({"num": num, "product": product_name, "qty": qty, "warehouse": warehouse})

    warehouses = list(dict.fromkeys(item["warehouse"] for item in data))

    ws.delete_rows(1, ws.max_row)
    ws.merged_cells.ranges = []
    ws.row_breaks = type(ws.row_breaks)()

    current_row = 1

    for wh in warehouses:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        cell = ws.cell(current_row, 1, value=azs_name)
        cell.font      = title_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        for col in range(1, 4):
            ws.cell(current_row, col).border = border
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        cell = ws.cell(current_row, 1, value=str(wh).upper())
        cell.font      = warehouse_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.fill      = gray_fill
        for col in range(1, 4):
            ws.cell(current_row, col).border = border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        for col, header in enumerate(["№", "ТОВАР", "КОЛ-ВО"], start=1):
            cell = ws.cell(current_row, col, value=header)
            cell.font      = header_font
            cell.alignment = center
            cell.fill      = gray_fill
            cell.border    = border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        for item in data:
            if item["warehouse"] == wh:
                ws.cell(current_row, 1, value=item["num"])
                ws.cell(current_row, 2, value=item["product"])
                ws.cell(current_row, 3, value=item["qty"])
                for col in range(1, 4):
                    cell = ws.cell(current_row, col)
                    cell.font      = text_font
                    cell.alignment = left
                    cell.border    = border
                current_row += 1

        ws.row_breaks.append(Break(id=current_row))
        current_row += 2

    for col_num, width in {1: 10, 2: 70, 3: 15}.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width

    ws.page_setup.orientation   = "portrait"
    ws.page_setup.fitToWidth    = 1
    ws.page_setup.fitToHeight   = False
    ws.sheet_view.showGridLines = False


if processed_sheets == 0:
    print("\n[ВНИМАНИЕ] Ни один лист основного файла не совпал с сетями из справочников.")
    print("Убедитесь, что названия листов содержат аббревиатуру сети (например: БП, МП, ПН).")


# =====================================================
# СОХРАНЕНИЕ
# =====================================================

base_name   = os.path.splitext(main_file)[0]
output_file = f"{base_name}_отсортированный.xlsm"

if os.path.exists(output_file):
    os.remove(output_file)

wb.save(output_file)


# =====================================================
# ЛОГ: не найденные товары
# =====================================================

if not_found:
    print("\n[ВНИМАНИЕ] Не найдены следующие позиции в справочнике:\n")
    grouped_errors = defaultdict(set)
    for sheet, item in not_found:
        grouped_errors[sheet].add(item)
    for i, sheet_name in enumerate(sorted(grouped_errors.keys()), start=1):
        print(f"{i}. Лист: \"{sheet_name}\"")
        for item in sorted(grouped_errors[sheet_name]):
            print(f"   ▪ {item}")
        print()
else:
    print("\n[ОК] Все товары успешно сопоставлены со справочником.")

print("Обработка завершена!")
print(f"Результат сохранён в файл:\n{output_file}")
